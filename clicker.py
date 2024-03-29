import pyautogui as gui
import logging
import time
import openpyxl

current_milli_time = lambda: int(round(time.time() * 1000))

def setup_logger(location):
    # logging                                                  
    logging.basicConfig(filename=location, filemode="w", level=logging.DEBUG)  

    # console handler  
    console = logging.StreamHandler()  
    console.setLevel(logging.ERROR)  
    logging.getLogger("").addHandler(console)

def load_workbook(location, maxRows):
    wb = openpyxl.load_workbook(location)
    ws = wb['Sheet1']

    db = []
    not_added = []

    for rownum in range(1,maxRows+1):
        try:
            hyperlink = ws.cell(row=rownum, column=1).hyperlink.target
            quantity = ws.cell(row=rownum, column=2).value

            if quantity != None:
                db.append([hyperlink, int(quantity)])
            else:
                not_added.append([hyperlink, quantity])
        except:
            logging.error("Did not add " + str(rownum))
    
    return db

def shop(db):
    for item in db:
        itemURL = item[0]
        quantity = item[1]
        failed = False

        #load URL in Chrome
        load_url(itemURL, quantity)

        #add to cart
        failed = click_add(itemURL, quantity, failed)
            
        #update quantity
        if quantity > 1:
            failed = click_plus(quantity, itemURL, failed)

def load_url(itemURL, quantity):
    gui.tripleClick(x=300, y=100)
    gui.typewrite(itemURL)
    gui.hotkey('enter')
    time.sleep(3)

def click_add(itemURL, quantity, failed):
    if failed == True:
        return True

    try:
        x1, y1 = gui.locateCenterOnScreen("add.png", grayscale=False, confidence=.5, region=(500,500, 1400,1000))
        gui.click(x=x1, y=y1, button='left')
    except:
        logging.error("Add-To-Cart-Missing: \n" + itemURL + "\n" + str(quantity) )
        return True

def click_plus(quantity, itemURL, failed):
    if failed == True:
        return True

    try:
        x1, y1 = gui.locateCenterOnScreen("plus.png", grayscale=False, confidence=.7, region=(0,0, 1500,1000))
        gui.click(x=x1, y=y1, button='left', clicks=(quantity-1), interval=.125)
    except:
        logging.error("Plus-Button-Missing: \n" + itemURL + "\n" + str(quantity) )
        return True

    try:
        gui.moveRel(100,100, duration=.125)
        gui.leftClick()
        x1, y1 = gui.locateCenterOnScreen("plus.png", grayscale=False, confidence=.7, region=(0,0, 1500,1000))
    except:
        logging.error("Possible-Quantity-Wrong: \n" + itemURL + "\n" + str(quantity) )
        return True


setup_logger("error.log")
maxRow = 83
db = load_workbook("Second Grocery Run.xlsx", maxRow)

#test DB
#db = [["https://grocery.walmart.com/ip/Great-Value-Deli-Style-Sliced-Pepper-Jack-Cheese-8-oz-12-count/155066735" , 13],["https://grocery.walmart.com/ip/Sam-s-Choice-Sliced-Monterey-Jack-Cheese-with-Jalapenos-Habaneros-and-Ghost-Peppers-8-Oz/146364618", 4],["https://grocery.walmart.com/ip/Great-Value-Cracker-Cuts-Pepper-Jack-Cheese-Slices-10-Oz/681794103" , 3],["https://grocery.walmart.com/ip/Wonder-Classic-White-Bread-20-oz-Loaf/37858875", 5]]
shop(db)